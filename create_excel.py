import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill('solid', fgColor='003366')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT = Font(name='맑은 고딕', size=10)
CELL_ALIGN = Alignment(vertical='center', wrap_text=True)
NUM_ALIGN = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
IDM_FILL = PatternFill('solid', fgColor='FFF2CC')
SELL_FILL = PatternFill('solid', fgColor='E2EFDA')
BEST_FILL = PatternFill('solid', fgColor='FCE4EC')
WEB_FILLS = [PatternFill('solid', fgColor='DAEEF3'), PatternFill('solid', fgColor='D6E4F0'), PatternFill('solid', fgColor='CDD8E6')]

def p15(cost):
    if not isinstance(cost, (int, float)): return '-'
    return math.ceil(cost * 1.15 / 100) * 100

def recommend(idm_sell, prices):
    """웹최저가와 아이디마켓 판매가 사이 적정단가 (중간값 100원 올림)"""
    nums = [p for p in prices if isinstance(p, (int, float))]
    idm_ok = isinstance(idm_sell, (int, float))
    if not nums and not idm_ok:
        return '-', ''
    if not nums and idm_ok:
        return idm_sell, '아이디마켓 기준'
    web_min = min(nums)
    if not idm_ok:
        return web_min, f'웹최저 {web_min:,}'
    if idm_sell <= web_min:
        return idm_sell, f'아이디마켓 유리 (웹최저 {web_min:,})'
    mid = math.ceil((idm_sell + web_min) / 2 / 100) * 100
    return mid, f'적정가 (IDM {idm_sell:,} ~ 웹 {web_min:,})'

def style_sheet(ws, headers, data, col_widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for r, rd in enumerate(data, 2):
        for c, val in enumerate(rd, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)].width = w

def fmt_and_color(ws, money_cols, color_map=None):
    for col in money_cols:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = NUM_ALIGN
                    if color_map and col in color_map:
                        cell.fill = color_map[col]

# =====================================================================
# 열 구조 (17열):
# A:No B:제조사 C:제품명 D:규격 E:단위
# F:아이디마켓원가 G:아이디마켓판매가(+15%)
# H:웹검색1 단가 I:웹검색1 판매처
# J:웹검색2 단가 K:웹검색2 판매처
# L:웹검색3 단가 M:웹검색3 판매처
# N:추천단가 O:산출근거 P:비고
# =====================================================================

H_COMMON = ['No','제조사','제품명','규격','단위',
    '아이디마켓\n원가(원)','아이디마켓\n판매가(원)\n(+15%)',
    '웹검색①\n단가(원)','웹검색①\n판매처',
    '웹검색②\n단가(원)','웹검색②\n판매처',
    '웹검색③\n단가(원)','웹검색③\n판매처',
    '추천단가\n(원)','산출근거','비고']
CW = [5,12,34,18,6, 14,14, 12,14, 12,14, 12,14, 14,30,18]
MONEY = [6,7,8,10,12,14]
CMAP = {6:IDM_FILL, 7:SELL_FILL, 8:WEB_FILLS[0], 10:WEB_FILLS[1], 12:WEB_FILLS[2], 14:BEST_FILL}

# raw: [No,제조사,제품명,규격,단위, idm원가, (w1가격,w1처),(w2가격,w2처),(w3가격,w3처), 비고]
def build_rows(raw_data):
    rows = []
    for r in raw_data:
        no,mk,nm,sp,un, idm_cost, w1p,w1s, w2p,w2s, w3p,w3s, note = r
        sell = p15(idm_cost)
        rec, basis = recommend(sell, [w1p, w2p, w3p])
        rows.append([no,mk,nm,sp,un, idm_cost,sell, w1p,w1s, w2p,w2s, w3p,w3s, rec,basis,note])
    return rows

# ========== 시트1: 에폭시 바닥재 ==========
ws1 = wb.active
ws1.title = '에폭시 바닥재'

raw1 = [
    # KCC 에폭시
    [1,'KCC','유니폭시 하도(ECO)','14L (2액형)','통', 64000, 74000,'11번가', 85000,'태산스토어', 105000,'KCC플러스', '콘크리트 바닥 하도'],
    [2,'KCC','유니폭시 하도 (일반형)','14L (2액형)','통', 60000, 78800,'11번가', '-','-', '-','-', 'ECO 아닌 일반형'],
    [3,'KCC','유니폭시코팅K 상도(녹색/회색)','16kg (2액형)','통', 76000, 85000,'태산스토어', 86900,'11번가', 105000,'11번가(타판매자)', '박막형 에폭시'],
    [4,'KCC','유니폭시코팅 상도(적갈색ECO)','16kg (2액형)','통', '-', 111900,'11번가', '-','-', '-','-', ''],
    [5,'KCC','유니폭시코팅 상도(백색ECO)','16kg (2액형)','통', '-', 116900,'11번가', '-','-', '-','-', ''],
    [6,'KCC','유니폭시코팅 상도(황색ECO)','16kg (2액형)','통', '-', 136900,'11번가', '-','-', '-','-', ''],
    [7,'KCC','유니폭시코팅 상도(투명)','12.6kg (2액형)','통', '-', 111900,'11번가', '-','-', '-','-', ''],
    [8,'KCC','유니폭시 라이닝(L)(후막형)','24kg (2액형)','통', 77000, 148000,'11번가', '-','글로벌MRO', '-','-', '도막3mm 무용제'],
    [9,'KCC','유니폭시 퍼티(빠데)','4L/9L (2액형)','통', '-', '-','-', '-','-', '-','-', '크랙 보수용'],
    [10,'KCC','숲으로 수니폭시 하도(수성)','16L','통', '-', 165000,'11번가', '-','-', '-','-', '친환경 수성'],
    # KCC 수성도료 (아이디마켓 전용)
    [11,'KCC','숲으로외부용KS1급','18L','통', 56000, '-','-', '-','-', '-','-', '수성'],
    [12,'KCC','숲으로내부용KS1급','18L','통', 47000, '-','-', '-','-', '-','-', '수성'],
    [13,'KCC','숲으로외부용플러스','18L','통', 48000, '-','-', '-','-', '-','-', '수성'],
    [14,'KCC','숲으로내부용플러스','18L','통', 44000, '-','-', '-','-', '-','-', '수성'],
    [15,'KCC','숲으로광택플러스','18L','통', 80000, '-','-', '-','-', '-','-', '수성'],
    # KCC 내화도료
    [16,'KCC','FIREMASK SQ-1003','18L','통', 95000, '-','-', '-','-', '-','-', '내화도료'],
    [17,'KCC','FIREMASK SQ-2003(기둥)/2007(보)','18L','통', 95000, '-','-', '-','-', '-','-', '내화도료'],
    # 노루 에폭시
    [18,'노루','크린폭시 하도','14L (2액형)','통', '-', 75000,'11번가', 91000,'11번가(타)', 98000,'페인트플러스', '콘크리트 바닥 하도'],
    [19,'노루','크린폭시 코팅플러스 상도','16kg (2액형)','통', '-', 98000,'11번가', 99000,'11번가(타)', 111000,'11번가(타)', '박막형 에폭시'],
    [20,'노루','크린폭시 라이닝(KS)','24kg','통', '-', 91000,'11번가', 116000,'11번가(타)', 127000,'11번가(타)', '무용제 후막형'],
    [21,'노루','크린폭시 라이닝 투명','18kg','통', '-', 206940,'스피드몰', '-','-', '-','-', '바닥 광택/방진'],
    [22,'노루','크린폭시 퍼티(속건형)','6kg (2액형)','통', '-', '-','-', '-','-', '-','-', '크랙 보수용'],
    [23,'노루','뉴워터폭시(수성에폭시)','16kg (2액형)','통', '-', '-','-', '-','-', '-','-', '수성 상도'],
]

d1 = build_rows(raw1)
style_sheet(ws1, H_COMMON, d1, CW)
fmt_and_color(ws1, MONEY, CMAP)

# ========== 시트2: 우레탄 방수재 ==========
ws2 = wb.create_sheet('우레탄 방수재')

raw2 = [
    [1,'KCC','스포탄 하도(ECO)','14kg','통', 53000, 63620,'다나와', 65000,'11번가', 76000,'태산스토어', ''],
    [2,'KCC','스포탄 노출방수재K(2액형 중도)','20kg','통', 52000, '-','-', '-','-', '-','-', ''],
    [3,'KCC','스포탄 노출방수재(2액형 중도)','20kg','통', 49000, '-','-', '-','-', '-','-', ''],
    [4,'KCC','스포탄 상도(2액형)','14kg','통', 80000, 103900,'11번가(12.6L)', 117900,'11번가(14kg)', 114990,'나비엠알오', ''],
    [5,'KCC','스포탄 KS1류방수재','20kg','통', 56000, '-','-', '-','-', '-','-', ''],
    [6,'KCC','스포탄 KS비노출방수재','24kg','통', 52800, '-','-', '-','-', '-','-', ''],
    [7,'KCC','하이퍼우레아 PU295A','400kg','통', 2240000, '-','-', '-','-', '-','-', ''],
    [8,'KCC','하이퍼우레아 HB195A','400kg','통', 1920000, '-','-', '-','-', '-','-', ''],
    [9,'KCC','스포탄 고경질실러','20kg','통', 66000, '-','-', '-','-', '-','-', ''],
    [10,'KCC','스포탄 고경질바닥재','20kg','통', 70000, '-','-', '-','-', '-','-', ''],
    [11,'KCC','스포탄 고경질상도','16.5kg','통', 88000, '-','-', '-','-', '-','-', ''],
    [12,'KCC','스포탄 KS하이퍼플로어','20kg','통', 76000, '-','-', '-','-', '-','-', ''],
    [13,'KCC','모노탄 1액형 노출방수재(중도)','16kg','통', '-', '-','-', '-','-', '-','-', '1액형'],
    [14,'KCC','모노탄 상도','3.6L','통', '-', '-','-', '-','-', '-','-', ''],
    [15,'KCC','스포탄 희석제','별도문의','통', '-', 44000,'KCC플러스', '-','-', '-','-', ''],
    [16,'KCC','코레실','별도문의','통', '-', 47000,'KCC플러스', '-','-', '-','-', ''],
    [17,'KCC','우레탄 옥상방수 2mm 30평 세트','30평형','set', '-', 851400,'페인트사랑', '-','-', '-','-', '세트'],
    [18,'KCC','우레탄 옥상방수 2mm 20평 세트','20평형','set', '-', 610500,'페인트사랑', '-','-', '-','-', '세트'],
    [19,'KCC','우레탄 옥상방수 1mm 15평 세트','15평형','set', '-', 295900,'페인트사랑', '-','-', '-','-', '세트'],
]

d2 = build_rows(raw2)
style_sheet(ws2, H_COMMON, d2, CW)
fmt_and_color(ws2, MONEY, CMAP)

# ========== 시트3: 판매점 연락처 ==========
ws3 = wb.create_sheet('판매점 연락처')
h3 = ['No','판매점명','대표자','주소','전화번호','팩스','이메일','홈페이지','사업자등록번호','취급품목']
d3 = [
    [1,'아이디마켓','-','-','별도문의','-','-','https://idfenc.com','-','KCC 도료/바닥재'],
    [2,'KCC플러스(다우상사)','도규옥','충북 청주시 흥덕구 가로수로 935 1층 102호','043-232-7770','-','sharonto@naver.com','https://kccplus.com','527-15-00834','KCC 에폭시/우레탄'],
    [3,'몰딩닷컴(금강데코)','노유진','대전 대덕구 오정동 73-5','010-4450-2014','042-635-0057','mol_ding@naver.com','https://mol-ding.com','305-81-39409','건축내외장재'],
    [4,'태산스토어(유한회사 태산)','이현준','전북 전주시 덕진구 기린대로 1031','010-5825-0155','-','marine851@naver.com','https://taesanstore.com','403-81-82267','에폭시/타일부자재'],
    [5,'페인트플러스','-','-','별도문의','-','-','https://paintplus.co.kr','-','노루페인트 전문'],
    [6,'KCC 본사','-','-','A/S: 080-022-8200','-','-','https://www.kccworld.co.kr','-','KCC 전제품'],
    [7,'노루페인트 본사','-','경기 안양시 만안구 박달로 351','031-467-6114','-','-','https://www.noroopaint.com','-','노루 전제품'],
    [8,'페인트사랑','-','-','별도문의','-','-','https://www.paintsarang.co.kr','-','방수재 세트'],
    [9,'새현산업','-','-','별도문의','-','-','https://shcnc.kr','-','보수보강 자재'],
    [10,'스피드몰','-','-','별도문의','-','-','https://www.speedmall.co.kr','-','산업용 자재'],
]
style_sheet(ws3, h3, d3, [5,24,10,42,20,16,24,30,18,20])

# 범례 / 조사일자
for ws in [ws1, ws2, ws3]:
    r = ws.max_row + 2
    notes = [
        '※ 아이디마켓 원가: KCC→아이디마켓 공급가 (2026.02) / 판매가: 원가×1.15 (100원 올림)',
        '※ 추천단가: 아이디마켓 판매가와 웹 최저가의 중간값 (적정가). 아이디마켓이 더 낮으면 아이디마켓 기준.',
        '※ 웹검색 단가: 2025~2026년 온라인 쇼핑몰 검색 기준 (시점에 따라 변동 가능)',
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=r+i, column=1, value=n)
        c.font = Font(name='맑은 고딕', size=9, italic=True, color='666666')

output = r'C:\Users\note\클호그\도료류_단가조사_2026-2월_rev1.xlsx'
wb.save(output)
print(f'완료: {output}')
