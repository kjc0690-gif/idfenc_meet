import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

def create_tbm_excel():
    """TBM 회의록 양식 엑셀 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TBM 회의록"

    # 페이지 설정
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=10)

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 열 너비 설정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35

    row = 1

    # 제목
    ws.merge_cells(f'A{row}:D{row}')
    title = ws[f'A{row}']
    title.value = "□ TBM 회의록 양식 (Tool Box Meeting 회의록)"
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 1

    # 기본 정보 섹션
    ws.merge_cells(f'A{row}:D{row}')
    header = ws[f'A{row}']
    header.value = "■ 기본 정보"
    header.fill = header_fill
    header.font = header_font
    header.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    # TBM 일시
    ws[f'A{row}'].value = "TBM 일시"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].value = "년     월     일,     시     분 ~ 시     분"
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # 현장명
    ws[f'A{row}'].value = "현장명"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 22
    row += 1

    # 작업명
    ws[f'A{row}'].value = "작업명"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 22
    row += 1

    # 작업내용
    ws[f'A{row}'].value = "작업내용"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 40
    row += 1

    # TBM 장소
    ws[f'A{row}'].value = "TBM 장소"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 22
    row += 1

    # 위험성평가
    ws[f'A{row}'].value = "위험성평가\n실시여부"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].value = "□ 예     □ 아니오"
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 30
    row += 2

    # 잠재위험요인 섹션
    ws.merge_cells(f'A{row}:D{row}')
    header = ws[f'A{row}']
    header.value = "■ 잠재위험요인 및 대책"
    header.fill = header_fill
    header.font = header_font
    header.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    # 표 헤더
    ws[f'A{row}'].value = "번호"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'B{row}'].value = "잠재위험요인"
    ws[f'B{row}'].fill = section_fill
    ws[f'B{row}'].font = section_font
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'C{row}'].value = "번호"
    ws[f'C{row}'].fill = section_fill
    ws[f'C{row}'].font = section_font
    ws[f'C{row}'].border = border_thin
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'D{row}'].value = "대책 (제거→대체→통제 순서 고려)"
    ws[f'D{row}'].fill = section_fill
    ws[f'D{row}'].font = section_font
    ws[f'D{row}'].border = border_thin
    ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = 22
    row += 1

    # 위험요인 행 (3개)
    for i in range(1, 4):
        ws[f'A{row}'].value = f"①" if i == 1 else f"②" if i == 2 else f"③"
        ws[f'A{row}'].border = border_thin
        ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="top")

        ws[f'B{row}'].border = border_thin
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")

        ws[f'C{row}'].value = f"①" if i == 1 else f"②" if i == 2 else f"③"
        ws[f'C{row}'].border = border_thin
        ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="top")

        ws[f'D{row}'].border = border_thin
        ws[f'D{row}'].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[row].height = 50
        row += 1

    row += 1

    # 중점위험요인 섹션
    ws.merge_cells(f'A{row}:D{row}')
    header = ws[f'A{row}']
    header.value = "■ 중점위험요인"
    header.fill = header_fill
    header.font = header_font
    header.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws[f'A{row}'].value = "선 정"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 40
    row += 1

    ws[f'A{row}'].value = "대 책"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 40
    row += 2

    # TBM 리더 확인
    ws.merge_cells(f'A{row}:D{row}')
    header = ws[f'A{row}']
    header.value = "■ TBM 리더 확인"
    header.fill = header_fill
    header.font = header_font
    header.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws[f'A{row}'].value = "소속"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 22
    row += 1

    ws[f'A{row}'].value = "직책"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 22
    row += 1

    ws[f'A{row}'].value = "성명"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 22
    row += 1

    ws[f'A{row}'].value = "서명"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin

    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].border = border_thin
    ws.row_dimensions[row].height = 50
    row += 2

    # 작업 전 안전조치 확인
    ws.merge_cells(f'A{row}:D{row}')
    header = ws[f'A{row}']
    header.value = "■ 작업 전 안전조치 확인"
    header.fill = header_fill
    header.font = header_font
    header.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws[f'A{row}'].value = "안전조치 항목"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'B{row}'].value = "조치여부"
    ws[f'B{row}'].fill = section_fill
    ws[f'B{row}'].font = section_font
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'C{row}'].value = "미조치 사항"
    ws[f'C{row}'].fill = section_fill
    ws[f'C{row}'].font = section_font
    ws[f'C{row}'].border = border_thin
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'D{row}'].value = "조치 내용"
    ws[f'D{row}'].fill = section_fill
    ws[f'D{row}'].font = section_font
    ws[f'D{row}'].border = border_thin
    ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = 22
    row += 1

    # 안전조치 항목들
    items = [
        ("①", "상부 작업자 하부 추락 위험"),
        ("②", "자재 양중시 낙하위험 / 낙하물"),
        ("③", "전도 사고"),
    ]

    for num, item in items:
        ws[f'A{row}'].value = f"{num} {item}"
        ws[f'A{row}'].border = border_thin
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="center")

        ws[f'B{row}'].value = "□ 예  □ 아니오"
        ws[f'B{row}'].border = border_thin
        ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")

        ws[f'C{row}'].border = border_thin
        ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical="top")

        ws[f'D{row}'].border = border_thin
        ws[f'D{row}'].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[row].height = 35
        row += 1

    row += 1

    # 참석자 확인
    ws.merge_cells(f'A{row}:D{row}')
    header = ws[f'A{row}']
    header.value = "■ 참석자 확인"
    header.fill = header_fill
    header.font = header_font
    header.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws[f'A{row}'].value = "이름"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'B{row}'].value = "서명"
    ws[f'B{row}'].fill = section_fill
    ws[f'B{row}'].font = section_font
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'C{row}'].value = "이름"
    ws[f'C{row}'].fill = section_fill
    ws[f'C{row}'].font = section_font
    ws[f'C{row}'].border = border_thin
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'D{row}'].value = "서명"
    ws[f'D{row}'].fill = section_fill
    ws[f'D{row}'].font = section_font
    ws[f'D{row}'].border = border_thin
    ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = 22
    row += 1

    # 참석자 행 (5명)
    for i in range(5):
        ws[f'A{row}'].border = border_thin
        ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

        ws[f'B{row}'].border = border_thin
        ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")

        ws[f'C{row}'].border = border_thin
        ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")

        ws[f'D{row}'].border = border_thin
        ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 40
        row += 1

    # 파일 저장
    output_path = Path(r"C:\Users\note\클호그\TBM\TBM_회의록_양식.xlsx")
    wb.save(output_path)
    print("TBM meeting form created successfully!")

if __name__ == "__main__":
    create_tbm_excel()
