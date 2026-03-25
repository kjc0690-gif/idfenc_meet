import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from datetime import datetime, date
from tkcalendar import DateEntry
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from pathlib import Path
from PIL import Image
import os
import shutil
import json

# 템플릿 파일 (이 파일을 복사해서 사용)
TEMPLATE_FILE = "20260318/TBM_회의록_20260316.xlsx"

# 회사명 고정
COMPANY_NAME = "(주)아이디에프이앤씨"

# 입력 이력 저장 파일
HISTORY_FILE = "입력이력.json"

# 위험요인/대책 자동 생성 템플릿
RISK_TEMPLATES = {
    "높이": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "전기": [
        ("감전 위험", "절연 도구 사용, 누전차단기 확인, 접지 확인"),
        ("화재 위험", "화기 금지 구역 설정, 소화기 비치, 가연성 물질 제거"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "폐쇄": [
        ("산소 부족 위험", "환기 실시, 산소 농도 측정, 구조 계획 수립"),
        ("질식 위험", "산소 호흡기 준비, 감시원 배치, 응급 대응 체계"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "중장비": [
        ("협착 사고", "신호원 배치, 기계 상태 확인, 접근 금지 구역 설정"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "방수": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "도장": [
        ("화학물질 중독 위험", "마스크 착용, 통풍 환기, 보안경 필수"),
        ("화재/폭발 위험", "화기 금지, 정전기 제거, 환기 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "견출": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "철거": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "용접": [
        ("화재/폭발 위험", "화기 금지 구역 설정, 소화기 비치, 가연성 물질 제거"),
        ("화상 위험", "방열복 착용, 용접 차광막 설치, 소화기 비치"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "굴착": [
        ("토사 붕괴 위험", "흙막이 설치, 경사면 안정 확인, 지반 조사"),
        ("매설물 파손 위험", "매설물 탐사, 수동 굴착, 관계기관 협의"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "타일": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "배관": [
        ("협착 사고", "신호원 배치, 기계 상태 확인, 접근 금지 구역 설정"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
}

PRIORITY_TEMPLATES = {
    "높이": ("추락방지", "안전대 착용 및 안전로프 사전 설치"),
    "전기": ("감전방지", "절연 도구 사용 및 누전차단기 확인"),
    "폐쇄": ("질식방지", "환기 실시 및 산소 농도 측정"),
    "중장비": ("협착방지", "신호원 배치 및 접근 금지 구역 설정"),
    "방수": ("전도방지", "작업장 주위 불필요한 잔재물 정리"),
    "도장": ("중독방지", "마스크 착용 및 환기 실시"),
    "견출": ("전도방지", "작업장 주위 불필요한 잔재물 정리"),
    "철거": ("추락방지", "안전대 착용 및 안전로프 사전 설치"),
    "용접": ("화재방지", "소화기 비치 및 가연성 물질 제거"),
    "굴착": ("붕괴방지", "흙막이 설치 및 경사면 안정 확인"),
    "타일": ("전도방지", "작업장 주위 불필요한 잔재물 정리"),
    "배관": ("협착방지", "신호원 배치 및 접근 금지 구역 설정"),
}


class DatePickerDialog(tk.Toplevel):
    """날짜 선택 대화상자"""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("TBM 작성 날짜 선택")
        self.geometry("350x200")
        self.resizable(False, False)
        self.result = None

        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text="TBM 회의록 날짜를 선택하세요:", font=("맑은 고딕", 11)).pack(padx=15, pady=(20, 10))

        # 날짜 선택 위젯 (과거 날짜도 선택 가능)
        self.cal = DateEntry(self, width=20, font=("맑은 고딕", 12),
                            date_pattern="yyyy-mm-dd", locale="ko_KR",
                            mindate=date(2020, 1, 1), maxdate=date(2030, 12, 31))
        self.cal.pack(padx=15, pady=10)

        # 버튼
        bf = ttk.Frame(self)
        bf.pack(pady=15)
        ttk.Button(bf, text="오늘 날짜", command=self._use_today, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="선택한 날짜", command=self._ok, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="취소", command=self._cancel, width=8).pack(side=tk.LEFT, padx=5)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self._cancel())

        # 가운데 정렬
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

        self.wait_window(self)

    def _use_today(self):
        self.result = date.today()
        self.destroy()

    def _ok(self):
        self.result = self.cal.get_date()
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


class AutoCompleteDialog(tk.Toplevel):
    """자동완성 기능이 있는 입력 창 - 앞글자 필터링"""
    def __init__(self, parent, title, prompt, history_list=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("450x300")
        self.resizable(False, False)
        self.result = None
        self.history_list = history_list or []

        self.transient(parent)
        self.grab_set()

        # 안내 문구
        ttk.Label(self, text=prompt, font=("맑은 고딕", 10), wraplength=400).pack(padx=15, pady=(15, 5))

        # 입력칸
        self.entry_var = tk.StringVar()
        self.entry_var.trace_add("write", self._on_entry_change)
        self.entry = ttk.Entry(self, textvariable=self.entry_var, font=("맑은 고딕", 11), width=40)
        self.entry.pack(padx=15, pady=5)
        self.entry.focus_set()

        # 이전 입력 목록 (자동완성)
        if self.history_list:
            ttk.Label(self, text="이전 입력값 (앞글자 입력 시 필터링, 더블클릭으로 선택):",
                      font=("맑은 고딕", 9)).pack(padx=15, pady=(10, 2), anchor="w")
        self.listbox = tk.Listbox(self, font=("맑은 고딕", 10), height=8)
        self.listbox.pack(padx=15, pady=2, fill=tk.BOTH, expand=True)
        self.listbox.bind("<<ListboxSelect>>", self._on_select)
        self.listbox.bind("<Double-Button-1>", self._on_double_click)

        # 목록 채우기
        self._update_listbox("")

        # 버튼
        bf = ttk.Frame(self)
        bf.pack(pady=10)
        ttk.Button(bf, text="확인", command=self._ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="취소", command=self._cancel, width=10).pack(side=tk.LEFT, padx=5)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self._cancel())

        # 창을 화면 가운데로
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

        self.wait_window(self)

    def _update_listbox(self, typed):
        self.listbox.delete(0, tk.END)
        typed_lower = typed.lower().strip()
        for item in self.history_list:
            if not typed_lower or item.lower().startswith(typed_lower) or typed_lower in item.lower():
                self.listbox.insert(tk.END, item)

    def _on_entry_change(self, *args):
        typed = self.entry_var.get()
        self._update_listbox(typed)
        # 자동 선택: 필터 결과가 1개면 자동으로 입력칸에 표시
        if self.listbox.size() == 1:
            self.listbox.selection_set(0)

    def _on_select(self, event):
        sel = self.listbox.curselection()
        if sel:
            self.entry_var.set(self.listbox.get(sel[0]))
            self.entry.icursor(tk.END)

    def _on_double_click(self, event):
        sel = self.listbox.curselection()
        if sel:
            self.entry_var.set(self.listbox.get(sel[0]))
            self._ok()

    def _ok(self):
        self.result = self.entry_var.get().strip()
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


class TBMRecorder:
    def __init__(self, root):
        self.root = root
        self.root.title("안전 TBM 회의록")
        self.root.geometry("500x220")
        self.base_dir = Path(__file__).resolve().parent
        self.template_path = self.base_dir / TEMPLATE_FILE

        # 회의록 저장 폴더
        self.save_dir = self.base_dir / "회의록"
        self.save_dir.mkdir(parents=True, exist_ok=True)

        # 입력 이력 로드
        self.history_path = self.base_dir / HISTORY_FILE
        self.history = self._load_history()

        self.selected_date = None  # 사용자가 선택한 날짜
        self.attendees = []
        self.current_site = ""
        self.work_name = ""
        self.work_content = ""
        self.tbm_location = ""
        self.company = COMPANY_NAME
        self.risk_factors = []
        self.risk_measures = []
        self.priority_risk = ""
        self.priority_measure = ""
        self.image_files = []

        self._create_main_ui()

    def _load_history(self):
        """이전 입력 이력 불러오기"""
        default = {"참석자": [], "현장명": [], "작업명": [], "TBM장소": [], "작업내용": [], "사진폴더": ""}
        try:
            if self.history_path.exists():
                with open(self.history_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for key in default:
                    if key not in data:
                        data[key] = default[key]
                return data
        except:
            pass
        return default

    def _save_history(self):
        """입력 이력 저장"""
        try:
            with open(self.history_path, "w", encoding="utf-8") as f:
                json.dump(self.history, f, ensure_ascii=False, indent=2)
        except:
            pass

    def _add_history(self, key, value):
        """이력에 값 추가 (중복 제거, 최근 것이 위로)"""
        if isinstance(self.history[key], list):
            if value in self.history[key]:
                self.history[key].remove(value)
            self.history[key].insert(0, value)
            self.history[key] = self.history[key][:50]

    def _create_main_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="안전 TBM 회의록", font=("맑은 고딕", 16, "bold")).pack(pady=10)

        today_str = datetime.now().strftime("%Y-%m-%d")
        ttk.Label(main_frame, text=f"오늘 날짜: {today_str}  (다른 날짜도 선택 가능)",
                  font=("맑은 고딕", 10)).pack(pady=5)

        bf = ttk.Frame(main_frame)
        bf.pack(pady=20)
        ttk.Button(bf, text="TBM 기록 시작", command=self._start_process, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="회의록 폴더 열기", command=self._open_folder, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="종료", command=self.root.quit, width=15).pack(side=tk.LEFT, padx=5)

    def _start_process(self):
        # 0. 날짜 선택
        date_dlg = DatePickerDialog(self.root)
        if date_dlg.result is None:
            return
        self.selected_date = date_dlg.result

        # 1. 참석자
        dlg = AutoCompleteDialog(self.root,
            "1. 참석자 입력",
            "참석자 이름을 입력하세요 (쉼표로 구분)\n예: 박상희, 이진규, 양종훈",
            self.history.get("참석자", []))
        if not dlg.result:
            return
        self.attendees = [a.strip() for a in dlg.result.split(',') if a.strip()]
        self._add_history("참석자", dlg.result)

        # 2. 현장명
        dlg = AutoCompleteDialog(self.root,
            "2. 현장명 입력",
            "현장명을 입력하세요\n예: 농협은행 울산영업부",
            self.history.get("현장명", []))
        if not dlg.result:
            return
        self.current_site = dlg.result
        self._add_history("현장명", dlg.result)

        # 3. 회사명은 고정 (입력 안 받음)

        # 4. 작업명
        dlg = AutoCompleteDialog(self.root,
            "3. 작업명 입력",
            "작업명을 입력하세요\n예: 습식방수공사",
            self.history.get("작업명", []))
        if not dlg.result:
            return
        self.work_name = dlg.result
        self._add_history("작업명", dlg.result)

        # 5. TBM 장소
        dlg = AutoCompleteDialog(self.root,
            "4. TBM 장소 입력",
            "TBM 장소를 입력하세요\n예: 지상1층",
            self.history.get("TBM장소", []))
        if not dlg.result:
            return
        self.tbm_location = dlg.result
        self._add_history("TBM장소", dlg.result)

        # 6. 작업내용
        w = tk.Toplevel(self.root)
        w.title("5. 작업내용 입력")
        w.geometry("500x400")
        ttk.Label(w, text="작업내용을 입력하세요:", font=("맑은 고딕", 11)).pack(padx=10, pady=(10, 5))

        # 이전 작업내용 목록
        prev_contents = self.history.get("작업내용", [])
        if prev_contents:
            ttk.Label(w, text="이전에 입력한 작업내용 (클릭하면 입력됩니다):", font=("맑은 고딕", 9)).pack(padx=10, anchor="w")
            lb = tk.Listbox(w, font=("맑은 고딕", 9), height=4)
            lb.pack(padx=10, pady=2, fill=tk.X)
            for item in prev_contents[:10]:
                short = item[:60] + "..." if len(item) > 60 else item
                lb.insert(tk.END, short)

        txt = scrolledtext.ScrolledText(w, height=8, width=60, font=("맑은 고딕", 10))
        txt.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

        if prev_contents:
            def on_lb_select(event):
                sel = lb.curselection()
                if sel:
                    txt.delete("1.0", tk.END)
                    txt.insert("1.0", prev_contents[sel[0]])
            lb.bind("<<ListboxSelect>>", on_lb_select)

        def next_step():
            self.work_content = txt.get("1.0", tk.END).strip()
            if not self.work_content:
                messagebox.showwarning("경고", "작업내용을 입력하세요!")
                return
            self._add_history("작업내용", self.work_content)
            w.destroy()
            self._auto_generate()

        ttk.Button(w, text="다음", command=next_step).pack(pady=10)
        w.transient(self.root)
        w.grab_set()

    def _auto_generate(self):
        """작업명+작업내용으로 위험요인/대책/중점위험요인 자동 생성"""
        search_text = (self.work_name + " " + self.work_content).lower()
        self.risk_factors = []
        self.risk_measures = []

        matched = False
        for key, risks in RISK_TEMPLATES.items():
            if key in search_text:
                for risk, measure in risks:
                    self.risk_factors.append(risk)
                    self.risk_measures.append(measure)
                matched = True
                break

        if not matched:
            self.risk_factors = ["상부 작업자 하부 추락 위험", "자재 양중시 낙하위험", "전도 사고"]
            self.risk_measures = [
                "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시",
                "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시",
                "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"
            ]

        # 중점위험요인
        matched_p = False
        for key, (p_risk, p_measure) in PRIORITY_TEMPLATES.items():
            if key in search_text:
                self.priority_risk = p_risk
                self.priority_measure = p_measure
                matched_p = True
                break
        if not matched_p:
            self.priority_risk = "전도방지"
            self.priority_measure = "작업장 주위 불필요한 잔재물 정리"

        # 사진 폴더 선택 (이전에 사용한 폴더가 있으면 자동 사용)
        last_folder = self.history.get("사진폴더", "")
        if last_folder and Path(last_folder).exists():
            result = messagebox.askyesnocancel("사진 폴더",
                f"이전에 사용한 사진 폴더를 사용할까요?\n\n{last_folder}\n\n"
                f"예 = 이 폴더 사용\n아니오 = 다른 폴더 선택\n취소 = 사진 없이 진행",
                parent=self.root)
            if result is True:
                folder = last_folder
            elif result is False:
                folder = filedialog.askdirectory(
                    title="사진 폴더를 선택하세요",
                    parent=self.root)
            else:
                folder = None
        else:
            folder = filedialog.askdirectory(
                title="사진 폴더를 선택하세요",
                parent=self.root)

        if folder:
            self.history["사진폴더"] = folder
            self._load_date_images(folder)

        self._show_summary()

    def _load_date_images(self, folder_path):
        """선택한 날짜의 사진 자동 선택"""
        self.image_files = []
        exts = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff'}
        target = self.selected_date
        target_str = target.strftime("%Y%m%d")
        target_str2 = target.strftime("%Y-%m-%d")
        target_str3 = target.strftime("%Y.%m.%d")
        target_short = target.strftime("%m%d")

        for f in sorted(Path(folder_path).iterdir()):
            if f.suffix.lower() not in exts:
                continue

            is_match = False
            fname = f.name

            # 파일명에 선택 날짜 포함?
            if target_str in fname or target_str2 in fname or target_str3 in fname or target_short in fname:
                is_match = True

            # 파일 수정일이 선택 날짜?
            if not is_match:
                mod_time = os.path.getmtime(f)
                mod_date = datetime.fromtimestamp(mod_time).date()
                if mod_date == target:
                    is_match = True

            # EXIF 촬영일이 선택 날짜?
            if not is_match:
                try:
                    img = Image.open(f)
                    exif_data = img._getexif()
                    if exif_data and 36867 in exif_data:
                        taken = exif_data[36867][:10].replace(":", "-")
                        if taken == target_str2:
                            is_match = True
                except:
                    pass

            if is_match:
                self.image_files.append(f)

        if self.image_files:
            messagebox.showinfo("사진 선택",
                f"{target_str2} 날짜 사진 {len(self.image_files)}장을 찾았습니다!",
                parent=self.root)
        else:
            messagebox.showinfo("사진 선택",
                f"{target_str2} 날짜 사진이 없습니다.\n사진 없이 진행합니다.",
                parent=self.root)

    def _show_summary(self):
        sw = tk.Toplevel(self.root)
        sw.title("TBM 기록 확인")
        sw.geometry("600x500")

        canvas = tk.Canvas(sw, bg="white")
        sb = ttk.Scrollbar(sw, orient="vertical", command=canvas.yview)
        sf = ttk.Frame(canvas)
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)

        leader = self.attendees[0] if self.attendees else ""
        date_str = self.selected_date.strftime("%Y-%m-%d")

        text = f"""
[TBM 회의록 미리보기]

날짜: {date_str}
현장명: {self.current_site}
소속: {self.company}
작업명: {self.work_name}
TBM장소: {self.tbm_location}

작업내용: {self.work_content}

잠재위험요인 / 대책:
"""
        nums = ["①", "②", "③"]
        for i, (r, m) in enumerate(zip(self.risk_factors, self.risk_measures)):
            text += f"  {nums[i]} {r}\n     → {m}\n"

        text += f"""
중점위험요인: {self.priority_risk}
중점대책: {self.priority_measure}

TBM 리더: {leader} (소속: {self.company})
참석자: {', '.join(self.attendees)} ({len(self.attendees)}명)
사진: {len(self.image_files)}장
"""
        ttk.Label(sf, text=text, justify=tk.LEFT, font=("맑은 고딕", 9)).pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        bf = ttk.Frame(sw)
        bf.pack(pady=10)
        ttk.Button(bf, text="저장", command=lambda: self._save_record(sw), width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="취소", command=sw.destroy, width=15).pack(side=tk.LEFT, padx=5)
        sw.transient(self.root)
        sw.grab_set()

    def _add_centered_image(self, ws, img_path, anchor_cell, max_w, max_h):
        """사진을 셀 영역 가운데에 정렬하여 삽입 (OneCellAnchor)"""
        try:
            img = XLImage(str(img_path))
            ratio = min(max_w / img.width, max_h / img.height)
            new_w = int(img.width * ratio)
            new_h = int(img.height * ratio)
            img.width = new_w
            img.height = new_h

            # 가운데 정렬 오프셋 (EMU 단위)
            offset_x = pixels_to_EMU((max_w - new_w) // 2)
            offset_y = pixels_to_EMU((max_h - new_h) // 2)

            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_letter, row_num = coordinate_from_string(anchor_cell)
            col_idx = column_index_from_string(col_letter) - 1

            marker = AnchorMarker(col=col_idx, colOff=offset_x, row=row_num - 1, rowOff=offset_y)
            size = XDRPositiveSize2D(pixels_to_EMU(new_w), pixels_to_EMU(new_h))
            anchor = OneCellAnchor(_from=marker, ext=size)
            img.anchor = anchor

            ws.add_image(img)
            return True
        except Exception as e:
            print(f"사진 오류 ({anchor_cell}): {e}")
            return False

    def _save_record(self, summary_window):
        try:
            dt = self.selected_date
            now_time = datetime.now()
            weekdays = ["월", "화", "수", "목", "금", "토", "일"]
            day_name = weekdays[dt.weekday()]

            # 템플릿 복사
            date_str = dt.strftime('%Y%m%d')
            file_name = f"TBM_회의록_{date_str}.xlsx"
            save_path = self.save_dir / file_name

            # 같은 날짜 파일이 이미 있으면 번호 추가
            counter = 1
            while save_path.exists():
                counter += 1
                file_name = f"TBM_회의록_{date_str}_{counter}.xlsx"
                save_path = self.save_dir / file_name

            shutil.copy2(self.template_path, save_path)

            # 복사한 파일 열기
            wb = openpyxl.load_workbook(save_path)

            # === 1페이지: TBM 회의록 ===
            ws = wb.worksheets[0]
            leader = self.attendees[0] if self.attendees else ""
            time_str = "08:00"

            # B3: TBM 일시
            ws['B3'] = f"{dt.year} 년 {dt.month:02d} 월 {dt.day:02d} 일, {time_str}~     작업날짜와 동일함 (예→■, 아니오□)"

            # B4: 작업명
            ws['B4'] = f"현장명 : {self.current_site}    작업명 : {self.work_name}"

            # B5: 작업내용
            ws['B5'] = self.work_content

            # B6: TBM장소
            ws['B6'] = self.tbm_location

            # A8~A10: 위험요인
            nums = ["①", "②", "③"]
            risk_cells = ['A8', 'A9', 'A10']
            measure_cells = ['D8', 'D9', 'D10']
            for i in range(3):
                risk = self.risk_factors[i] if i < len(self.risk_factors) else ""
                measure = self.risk_measures[i] if i < len(self.risk_measures) else ""
                ws[risk_cells[i]] = f"{nums[i]} {risk}"
                ws[measure_cells[i]] = f"{nums[i]} {measure}"

            # C11: 중점위험요인 선정
            ws['C11'] = self.priority_risk
            # C12: 중점위험요인 대책
            ws['C12'] = self.priority_measure

            # A13: TBM 리더확인
            ws['A13'] = f"TBM 리더확인  · 소속 {self.company}  · 직책 반장  · 성명 {leader} (서명)"

            # A16~A18: 안전조치 확인 (위험요인 반복)
            safety_cells = ['A16', 'A17', 'A18']
            for i in range(3):
                risk = self.risk_factors[i] if i < len(self.risk_factors) else ""
                ws[safety_cells[i]] = f"{nums[i]} {risk}"

            # 참석자 이름 (A24~, 3열 배치)
            for r in range(24, 30):
                for c in ['A', 'C', 'E']:
                    ws[f'{c}{r}'] = ""

            col_pairs = ['A', 'C', 'E']
            max_per_col = max((len(self.attendees) + 2) // 3, 1)
            for idx, name in enumerate(self.attendees):
                col_idx = idx // max_per_col
                row_idx = idx % max_per_col
                if col_idx < 3:
                    cell = f'{col_pairs[col_idx]}{24 + row_idx}'
                    ws[cell] = name

            # === 2페이지: 사진대지 ===
            ws2 = wb.worksheets[1]
            ws2._images = []

            ws2['A2'] = f"현장명 : {self.current_site}"
            ws2['A3'] = f"작업명 : {self.company} ({self.work_name})"
            ws2['A4'] = f"날   짜 : {dt.year}. {dt.month:02d}. {dt.day:02d} ({day_name})"
            ws2['B6'] = self.work_name
            ws2['B7'] = self.tbm_location

            # 사진 넣기 (가운데 정렬)
            if self.image_files:
                # 첫 번째 사진 (큰 사진 - A5 영역)
                self._add_centered_image(ws2, self.image_files[0], 'A5', 500, 380)

                # 두 번째 사진 (왼쪽 아래 - A8 영역)
                if len(self.image_files) > 1:
                    self._add_centered_image(ws2, self.image_files[1], 'A8', 240, 250)

                # 세 번째 사진 (오른쪽 아래 - C8 영역)
                if len(self.image_files) > 2:
                    self._add_centered_image(ws2, self.image_files[2], 'C8', 240, 250)

            wb.save(save_path)

            # 이력 저장
            self._save_history()

            messagebox.showinfo("저장 완료",
                f"TBM 회의록이 저장되었습니다!\n\n"
                f"날짜: {dt.strftime('%Y-%m-%d')}\n"
                f"파일: {file_name}\n"
                f"위치: 회의록 폴더",
                parent=self.root)
            summary_window.destroy()

            # 저장 후 파일 열기
            os.startfile(save_path)

        except Exception as e:
            messagebox.showerror("오류", f"저장 중 오류: {str(e)}", parent=self.root)

    def _open_folder(self):
        try:
            os.startfile(self.save_dir)
        except Exception as e:
            messagebox.showerror("오류", f"폴더를 열 수 없습니다: {str(e)}", parent=self.root)


if __name__ == "__main__":
    root = tk.Tk()
    app = TBMRecorder(root)
    root.mainloop()
