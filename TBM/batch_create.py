"""6일치 TBM 회의록 일괄 생성"""
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from datetime import datetime
import shutil

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE = BASE_DIR / "20260318" / "TBM_회의록_20260316.xlsx"
SAVE_DIR = BASE_DIR / "회의록"
SAVE_DIR.mkdir(parents=True, exist_ok=True)
PHOTO_DIR = BASE_DIR / "사진"

# 공통 정보
COMPANY = "(주)아이디에프이앤씨"
SITE = "온양농협 하나로마트 신축"
WORK_NAME = "습식방수공사"
TBM_LOCATION = "지상1층"
WORK_CONTENT = "우레탄 방수 시공"
ATTENDEES = ["최준호", "한웅철", "남만우"]

# 방수 키워드 매칭 위험요인
RISK_FACTORS = [
    "상부 작업자 하부 추락 위험",
    "자재 양중시 낙하위험",
    "전도 사고",
]
RISK_MEASURES = [
    "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시",
    "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시",
    "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판",
]
PRIORITY_RISK = "전도방지"
PRIORITY_MEASURE = "작업장 주위 불필요한 잔재물 정리"

# 날짜별 사진 매핑
DATES = [
    "20260316",
    "20260317",
    "20260318",
    "20260319",
    "20260320",
    "20260321",
]

WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]

for date_str in DATES:
    dt = datetime.strptime(date_str, "%Y%m%d")
    day_name = WEEKDAYS[dt.weekday()]
    leader = ATTENDEES[0]

    # 사진 찾기
    photo = None
    for ext in ['.jpg', '.jpeg', '.png']:
        p = PHOTO_DIR / f"{date_str}{ext}"
        if p.exists():
            photo = p
            break

    # 템플릿 복사
    file_name = f"TBM_회의록_{date_str}.xlsx"
    save_path = SAVE_DIR / file_name
    counter = 1
    while save_path.exists():
        counter += 1
        file_name = f"TBM_회의록_{date_str}_{counter}.xlsx"
        save_path = SAVE_DIR / file_name

    shutil.copy2(TEMPLATE, save_path)
    wb = openpyxl.load_workbook(save_path)

    # === 1페이지: TBM 회의록 ===
    ws = wb.worksheets[0]
    time_str = "08:00"

    ws['B3'] = f"{dt.year} 년 {dt.month:02d} 월 {dt.day:02d} 일, {time_str}~     작업날짜와 동일함 (예→■, 아니오□)"
    ws['B4'] = f"현장명 : {SITE}    작업명 : {WORK_NAME}"
    ws['B5'] = WORK_CONTENT
    ws['B6'] = TBM_LOCATION

    nums = ["①", "②", "③"]
    risk_cells = ['A8', 'A9', 'A10']
    measure_cells = ['D8', 'D9', 'D10']
    for i in range(3):
        ws[risk_cells[i]] = f"{nums[i]} {RISK_FACTORS[i]}"
        ws[measure_cells[i]] = f"{nums[i]} {RISK_MEASURES[i]}"

    ws['C11'] = PRIORITY_RISK
    ws['C12'] = PRIORITY_MEASURE
    ws['A13'] = f"TBM 리더확인  · 소속 {COMPANY}  · 직책 반장  · 성명 {leader} (서명)"

    safety_cells = ['A16', 'A17', 'A18']
    for i in range(3):
        ws[safety_cells[i]] = f"{nums[i]} {RISK_FACTORS[i]}"

    for r in range(24, 30):
        for c in ['A', 'C', 'E']:
            ws[f'{c}{r}'] = ""

    col_pairs = ['A', 'C', 'E']
    max_per_col = max((len(ATTENDEES) + 2) // 3, 1)
    for idx, name in enumerate(ATTENDEES):
        col_idx = idx // max_per_col
        row_idx = idx % max_per_col
        if col_idx < 3:
            ws[f'{col_pairs[col_idx]}{24 + row_idx}'] = name

    # === 2페이지: 사진대지 ===
    ws2 = wb.worksheets[1]
    ws2._images = []

    ws2['A2'] = f"현장명 : {SITE}"
    ws2['A3'] = f"작업명 : {COMPANY} ({WORK_NAME})"
    ws2['A4'] = f"날   짜 : {dt.year}. {dt.month:02d}. {dt.day:02d} ({day_name})"
    ws2['B6'] = WORK_NAME
    ws2['B7'] = TBM_LOCATION

    if photo:
        try:
            img = XLImage(str(photo))
            max_w, max_h = 500, 380
            ratio = min(max_w / img.width, max_h / img.height)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
            ws2.add_image(img, 'A5')
        except Exception as e:
            print(f"  사진 오류: {e}")

    wb.save(save_path)
    print(f"[완료] {file_name} ({dt.month}/{dt.day} {day_name}) - 사진: {'있음' if photo else '없음'}")

print(f"\n총 {len(DATES)}건 생성 완료! 저장 위치: {SAVE_DIR}")
