import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from PIL import Image
import os
import shutil

# 템플릿 파일 (이 파일을 복사해서 사용)
TEMPLATE_FILE = "20260318/TBM_회의록_20260316.xlsx"

# 위험요인/대책 자동 생성 템플릿
RISK_TEMPLATES = {
    "높이": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "전기": [
        ("감전 위험", "절연 도구 사용, 누전차단기 확인, 접지 확인"),
        ("화재 위험", "화기 금지 구역 설정, 소화기 비치, 가연성 물질 제거"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "폐쇄": [
        ("산소 부족 위험", "환기 실시, 산소 농도 측정, 구조 계획 수립"),
        ("질식 위험", "산소 호흡기 준비, 감시원 배치, 응급 대응 체계"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "중장비": [
        ("협착 사고", "신호원 배치, 기계 상태 확인, 접근 금지 구역 설정"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "방수": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "도장": [
        ("화학물질 중독 위험", "마스크 착용, 통풍 환기, 보안경 필수"),
        ("화재/폭발 위험", "화기 금지, 정전기 제거, 환기 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "견출": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "철거": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "용접": [
        ("화재/폭발 위험", "화기 금지 구역 설정, 소화기 비치, 가연성 물질 제거"),
        ("화상 위험", "방열복 착용, 용접 차광막 설치, 소화기 비치"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "굴착": [
        ("토사 붕괴 위험", "흙막이 설치, 경사면 안정 확인, 지반 조사"),
        ("매설물 파손 위험", "매설물 탐사, 수동 굴착, 관계기관 협의"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
    "타일": [
        ("상부 작업자 하부 추락 위험", "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지테이프, 미끄럼방지타일. 타공발판"),
    ],
    "배관": [
        ("협착 사고", "신호원 배치, 기계 상태 확인, 접근 금지 구역 설정"),
        ("자재 양중시 낙하위험", "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시"),
        ("전도 사고", "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"),
    ],
}

PRIORITY_TEMPLATES = {
    "높이": ("추락방지", "안전대 착용 및 안전로프 사전 설치"),
    "전기": ("감전방지", "절연 도구 사용 및 누전차단기 확인"),
    "폐쇄": ("질식방지", "환기 실시 및 산소 농도 측정"),
    "중장비": ("협착방지", "신호원 배치 및 접근 금지 구역 설정"),
    "방수": ("전도방지", "작업장 주위 불필요한 잔재물 정리"),
    "도장": ("중독방지", "마스크 착용 및 환기 실시"),
    "견출": ("전도방지", "작업장 주위 불필요한 잔재물 정리"),
    "철거": ("추락방지", "안전대 착용 및 안전로프 사전 설치"),
    "용접": ("화재방지", "소화기 비치 및 가연성 물질 제거"),
    "굴착": ("붕괴방지", "흙막이 설치 및 경사면 안정 확인"),
    "타일": ("전도방지", "작업장 주위 불필요한 잔재물 정리"),
    "배관": ("협착방지", "신호원 배치 및 접근 금지 구역 설정"),
}


class TBMRecorder:
    def __init__(self, root):
        self.root = root
        self.root.title("안전 TBM 회의록")
        self.root.geometry("500x200")
        self.base_dir = Path(__file__).resolve().parent
        self.template_path = self.base_dir / TEMPLATE_FILE

        # 회의록 저장 폴더
        self.save_dir = self.base_dir / "회의록"
        self.save_dir.mkdir(exist_ok=True)

        self.date = datetime.now().strftime("%Y-%m-%d")
        self.attendees = []
        self.current_site = ""
        self.work_name = ""
        self.work_content = ""
        self.tbm_location = ""
        self.company = ""
        self.risk_factors = []
        self.risk_measures = []
        self.priority_risk = ""
        self.priority_measure = ""
        self.image_files = []

        self._create_main_ui()

    def _create_main_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="안전 TBM 회의록", font=("맑은 고딕", 16, "bold")).pack(pady=10)
        ttk.Label(main_frame, text=f"오늘 날짜: {self.date}", font=("맑은 고딕", 11)).pack(pady=5)
        bf = ttk.Frame(main_frame)
        bf.pack(pady=20)
        ttk.Button(bf, text="TBM 기록 시작", command=self._start_process, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="회의록 폴더 열기", command=self._open_folder, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="종료", command=self.root.quit, width=15).pack(side=tk.LEFT, padx=5)

    def _start_process(self):
        # 1. 참석자
        inp = simpledialog.askstring("1. 참석자 입력",
            "참석자 이름을 입력하세요\n(쉼표로 구분)\n\n예: 최준호, 한웅철, 남만우",
            parent=self.root)
        if not inp:
            return
        self.attendees = [a.strip() for a in inp.split(',') if a.strip()]

        # 2. 현장명
        v = simpledialog.askstring("2. 현장명 입력",
            "현장명을 입력하세요\n\n예: 온양농협 대규모점포(하나로마트) 신축공사",
            parent=self.root)
        if not v:
            return
        self.current_site = v

        # 3. 소속(회사명)
        v = simpledialog.askstring("3. 소속 입력",
            "소속(회사명)을 입력하세요\n\n예: ㈜아이디에프이앤씨",
            parent=self.root)
        if not v:
            return
        self.company = v

        # 4. 작업명
        v = simpledialog.askstring("4. 작업명 입력",
            "작업명을 입력하세요\n\n예: 습식방수공사",
            parent=self.root)
        if not v:
            return
        self.work_name = v

        # 5. TBM 장소
        v = simpledialog.askstring("5. TBM 장소 입력",
            "TBM 장소를 입력하세요\n\n예: 지상1층",
            parent=self.root)
        if not v:
            return
        self.tbm_location = v

        # 6. 작업내용
        w = tk.Toplevel(self.root)
        w.title("6. 작업내용 입력")
        w.geometry("500x300")
        ttk.Label(w, text="작업내용을 입력하세요:", font=("맑은 고딕", 11)).pack(padx=10, pady=10)
        txt = scrolledtext.ScrolledText(w, height=10, width=60, font=("맑은 고딕", 10))
        txt.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

        def next_step():
            self.work_content = txt.get("1.0", tk.END).strip()
            if not self.work_content:
                messagebox.showwarning("경고", "작업내용을 입력하세요!")
                return
            w.destroy()
            self._auto_generate()

        ttk.Button(w, text="다음", command=next_step).pack(pady=10)
        w.transient(self.root)
        w.grab_set()

    def _auto_generate(self):
        """작업명+작업내용으로 위험요인/대책/중점위험요인 자동 생성"""
        search_text = (self.work_name + " " + self.work_content).lower()
        self.risk_factors = []
        self.risk_measures = []

        matched = False
        for key, risks in RISK_TEMPLATES.items():
            if key in search_text:
                for risk, measure in risks:
                    self.risk_factors.append(risk)
                    self.risk_measures.append(measure)
                matched = True
                break

        if not matched:
            self.risk_factors = ["상부 작업자 하부 추락 위험", "자재 양중시 낙하위험", "전도 사고"]
            self.risk_measures = [
                "비계 설치등 상부 작업자 안전대 착용 및 안전로프 사전 설치 후 실시",
                "자재양중작업 전 슬링벨트 상태 및 결손상태 확인후 양중 실시",
                "작업장 바닥정리 / 보호구 착용 / 미끄럼방지 조치"
            ]

        # 중점위험요인
        matched_p = False
        for key, (p_risk, p_measure) in PRIORITY_TEMPLATES.items():
            if key in search_text:
                self.priority_risk = p_risk
                self.priority_measure = p_measure
                matched_p = True
                break
        if not matched_p:
            self.priority_risk = "전도방지"
            self.priority_measure = "작업장 주위 불필요한 잔재물 정리"

        # 사진 폴더 선택
        folder = filedialog.askdirectory(
            title="사진 폴더를 선택하세요 (오늘 날짜 사진만 자동 선택됩니다)",
            parent=self.root)
        if folder:
            self._load_today_images(folder)

        self._show_summary()

    def _load_today_images(self, folder_path):
        """오늘 날짜 사진만 자동 선택"""
        self.image_files = []
        exts = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff'}
        today = datetime.now().date()
        today_str = today.strftime("%Y%m%d")
        today_str2 = today.strftime("%Y-%m-%d")
        today_str3 = today.strftime("%Y.%m.%d")
        today_short = today.strftime("%m%d")

        for f in sorted(Path(folder_path).iterdir()):
            if f.suffix.lower() not in exts:
                continue

            is_today = False
            fname = f.name

            # 파일명에 오늘 날짜 포함?
            if today_str in fname or today_str2 in fname or today_str3 in fname or today_short in fname:
                is_today = True

            # 파일 수정일이 오늘?
            if not is_today:
                mod_time = os.path.getmtime(f)
                mod_date = datetime.fromtimestamp(mod_time).date()
                if mod_date == today:
                    is_today = True

            # EXIF 촬영일이 오늘?
            if not is_today:
                try:
                    img = Image.open(f)
                    exif = img._getexif()
                    if exif and 36867 in exif:
                        taken = exif[36867][:10].replace(":", "-")
                        if taken == today_str2:
                            is_today = True
                except:
                    pass

            if is_today:
                self.image_files.append(f)

        if self.image_files:
            messagebox.showinfo("사진 선택",
                f"오늘 날짜 사진 {len(self.image_files)}장을 찾았습니다!",
                parent=self.root)
        else:
            result = messagebox.askyesno("사진 없음",
                "오늘 날짜 사진이 없습니다.\n폴더의 모든 사진을 사용할까요?",
                parent=self.root)
            if result:
                for f in sorted(Path(folder_path).iterdir()):
                    if f.suffix.lower() in exts:
                        self.image_files.append(f)

    def _show_summary(self):
        sw = tk.Toplevel(self.root)
        sw.title("TBM 기록 확인")
        sw.geometry("600x500")

        canvas = tk.Canvas(sw, bg="white")
        sb = ttk.Scrollbar(sw, orient="vertical", command=canvas.yview)
        sf = ttk.Frame(canvas)
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)

        leader = self.attendees[0] if self.attendees else ""

        text = f"""
[TBM 회의록 미리보기]

날짜: {self.date}
현장명: {self.current_site}
소속: {self.company}
작업명: {self.work_name}
TBM장소: {self.tbm_location}

작업내용: {self.work_content}

잠재위험요인 / 대책:
"""
        nums = ["①", "②", "③"]
        for i, (r, m) in enumerate(zip(self.risk_factors, self.risk_measures)):
            text += f"  {nums[i]} {r}\n     → {m}\n"

        text += f"""
중점위험요인: {self.priority_risk}
중점대책: {self.priority_measure}

TBM 리더: {leader} (소속: {self.company})
참석자: {', '.join(self.attendees)} ({len(self.attendees)}명)
사진: {len(self.image_files)}장
"""
        ttk.Label(sf, text=text, justify=tk.LEFT, font=("맑은 고딕", 9)).pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        bf = ttk.Frame(sw)
        bf.pack(pady=10)
        ttk.Button(bf, text="저장", command=lambda: self._save_record(sw), width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(bf, text="취소", command=sw.destroy, width=15).pack(side=tk.LEFT, padx=5)
        sw.transient(self.root)
        sw.grab_set()

    def _save_record(self, summary_window):
        try:
            now = datetime.now()
            weekdays = ["월", "화", "수", "목", "금", "토", "일"]
            day_name = weekdays[now.weekday()]

            # 템플릿 복사
            file_name = f"TBM_회의록_{now.strftime('%Y%m%d')}.xlsx"
            save_path = self.save_dir / file_name

            # 같은 날짜 파일이 이미 있으면 번호 추가
            counter = 1
            while save_path.exists():
                counter += 1
                file_name = f"TBM_회의록_{now.strftime('%Y%m%d')}_{counter}.xlsx"
                save_path = self.save_dir / file_name

            shutil.copy2(self.template_path, save_path)

            # 복사한 파일 열기
            wb = openpyxl.load_workbook(save_path)

            # === 1페이지: TBM 회의록 ===
            ws = wb.worksheets[0]
            leader = self.attendees[0] if self.attendees else ""
            time_str = now.strftime("%H:%M")

            # B3: TBM 일시
            ws['B3'] = f"{now.year} 년 {now.month:02d} 월 {now.day:02d} 일, {time_str}~     작업날짜와 동일함 (예→■, 아니오□)"

            # B4: 작업명
            ws['B4'] = f"현장명 : {self.current_site}    작업명 : {self.work_name}"

            # B5: 작업내용
            ws['B5'] = self.work_content

            # B6: TBM장소
            ws['B6'] = self.tbm_location

            # A8~A10: 위험요인
            nums = ["①", "②", "③"]
            risk_cells = ['A8', 'A9', 'A10']
            measure_cells = ['D8', 'D9', 'D10']
            for i in range(3):
                risk = self.risk_factors[i] if i < len(self.risk_factors) else ""
                measure = self.risk_measures[i] if i < len(self.risk_measures) else ""
                ws[risk_cells[i]] = f"{nums[i]} {risk}"
                ws[measure_cells[i]] = f"{nums[i]} {measure}"

            # C11: 중점위험요인 선정
            ws['C11'] = self.priority_risk
            # C12: 중점위험요인 대책
            ws['C12'] = self.priority_measure

            # A13: TBM 리더확인
            ws['A13'] = f"TBM 리더확인  · 소속 {self.company}  · 직책 반장  · 성명 {leader} (서명)"

            # A16~A18: 안전조치 확인 (위험요인 반복)
            safety_cells = ['A16', 'A17', 'A18']
            for i in range(3):
                risk = self.risk_factors[i] if i < len(self.risk_factors) else ""
                ws[safety_cells[i]] = f"{nums[i]} {risk}"

            # 참석자 이름 (A24~, 3열 배치)
            # 기존 참석자 이름 지우기
            for r in range(24, 30):
                for c in ['A', 'C', 'E']:
                    ws[f'{c}{r}'] = ""

            # 참석자 채우기 (세로 방향으로 3열 배치)
            col_pairs = ['A', 'C', 'E']
            max_per_col = max((len(self.attendees) + 2) // 3, 1)
            for idx, name in enumerate(self.attendees):
                col_idx = idx // max_per_col
                row_idx = idx % max_per_col
                if col_idx < 3:
                    cell = f'{col_pairs[col_idx]}{24 + row_idx}'
                    ws[cell] = name

            # === 2페이지: 사진대지 ===
            ws2 = wb.worksheets[1]

            # 기존 이미지 제거
            ws2._images = []

            # A2: 현장명
            ws2['A2'] = f"현장명 : {self.current_site}"
            # A3: 작업명
            ws2['A3'] = f"작업명 : {self.company} ({self.work_name})"
            # A4: 날짜
            ws2['A4'] = f"날   짜 : {now.year}. {now.month:02d}. {now.day:02d} ({day_name})"

            # B6: 공종
            ws2['B6'] = self.work_name
            # B7: 위치
            ws2['B7'] = self.tbm_location

            # 사진 넣기 (원본 유지)
            if self.image_files:
                # 첫 번째 사진 (큰 사진 - A5 영역)
                try:
                    img = XLImage(str(self.image_files[0]))
                    # 원본 비율 유지하면서 셀에 맞추기
                    max_w, max_h = 500, 380
                    ratio = min(max_w / img.width, max_h / img.height)
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                    ws2.add_image(img, 'A5')
                except Exception as e:
                    print(f"사진1 오류: {e}")

                # 두 번째 사진 (왼쪽 아래 - A8 영역)
                if len(self.image_files) > 1:
                    try:
                        img2 = XLImage(str(self.image_files[1]))
                        max_w, max_h = 240, 250
                        ratio = min(max_w / img2.width, max_h / img2.height)
                        img2.width = int(img2.width * ratio)
                        img2.height = int(img2.height * ratio)
                        ws2.add_image(img2, 'A8')
                    except Exception as e:
                        print(f"사진2 오류: {e}")

                # 세 번째 사진 (오른쪽 아래 - C8 영역)
                if len(self.image_files) > 2:
                    try:
                        img3 = XLImage(str(self.image_files[2]))
                        max_w, max_h = 240, 250
                        ratio = min(max_w / img3.width, max_h / img3.height)
                        img3.width = int(img3.width * ratio)
                        img3.height = int(img3.height * ratio)
                        ws2.add_image(img3, 'C8')
                    except Exception as e:
                        print(f"사진3 오류: {e}")

            wb.save(save_path)

            messagebox.showinfo("저장 완료",
                f"TBM 회의록이 저장되었습니다!\n\n"
                f"파일: {file_name}\n"
                f"위치: 회의록 폴더",
                parent=self.root)
            summary_window.destroy()

            # 저장 후 파일 열기
            os.startfile(save_path)

        except Exception as e:
            messagebox.showerror("오류", f"저장 중 오류: {str(e)}", parent=self.root)

    def _open_folder(self):
        try:
            os.startfile(self.save_dir)
        except Exception as e:
            messagebox.showerror("오류", f"폴더를 열 수 없습니다: {str(e)}", parent=self.root)


if __name__ == "__main__":
    root = tk.Tk()
    app = TBMRecorder(root)
    root.mainloop()
