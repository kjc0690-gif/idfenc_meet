"""
(주)아이디에프이앤씨 AS 대응 시스템 - 엑셀 양식 생성기

AS 프로세스에서 수집한 데이터를 엑셀 양식(공문접수서, 시공일지, 회의록)으로 변환한다.
JSON 파일 또는 딕셔너리를 입력받아 엑셀 파일을 생성한다.
"""

import json
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter


# ──── 공통 스타일 ────
THIN = Side(style='thin')
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FONT = Font(name='맑은 고딕', bold=True, size=11)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=16)
SUBTITLE_FONT = Font(name='맑은 고딕', bold=True, size=12)
NORMAL_FONT = Font(name='맑은 고딕', size=10)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT_W = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
LIGHT_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
WRAP = Alignment(wrap_text=True, vertical='center')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _apply_border(ws, row_start, row_end, col_start, col_end):
    """셀 범위에 테두리를 적용한다."""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL


def _set_cell(ws, row, col, value, font=None, alignment=None, fill=None):
    """셀에 값과 스타일을 설정한다."""
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if fill: cell.fill = fill
    cell.border = BORDER_ALL
    return cell


# ═══════════════════════════════════════════════
# 1. AS 공문접수서
# ═══════════════════════════════════════════════

def create_document_receipt(data: dict, output_path: str) -> str:
    """AS 공문접수서 엑셀을 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'AS 공문접수서'

    # 열 너비
    widths = [4, 14, 18, 14, 18, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    meta = data.get('metadata', {})
    basic = data.get('basicInfo', {})
    urgency = data.get('urgency', {})

    # 제목
    ws.merge_cells('A1:G1')
    _set_cell(ws, 1, 1, '(주)아이디에프이앤씨 AS 공문접수서', TITLE_FONT, CENTER)
    ws.row_dimensions[1].height = 40

    # 접수 정보
    row = 3
    _set_cell(ws, row, 1, 'No', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'B{row}:C{row}')
    _set_cell(ws, row, 2, meta.get('caseId', ''), NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '접수일', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:G{row}')
    created = meta.get('createdAt', '')
    if created:
        try:
            created = datetime.fromisoformat(created.replace('Z', '+00:00')).strftime('%Y-%m-%d')
        except:
            pass
    _set_cell(ws, row, 5, created, NORMAL_FONT, CENTER)

    row = 4
    _set_cell(ws, row, 1, '', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 2, '현장명', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'C{row}:G{row}')
    _set_cell(ws, row, 3, data.get('siteName', ''), NORMAL_FONT, WRAP)

    row = 5
    _set_cell(ws, row, 1, '', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 2, '현장개요', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'C{row}:G{row+1}')
    _set_cell(ws, row, 3, data.get('siteOverview', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row].height = 30
    ws.row_dimensions[row+1].height = 30
    # 빈 셀 테두리
    for c in range(1, 8):
        ws.cell(row=row+1, column=c).border = BORDER_ALL

    # 건물 정보
    row = 7
    ws.merge_cells(f'A{row}:G{row}')
    _set_cell(ws, row, 1, '■ 건물 정보', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    fields = [
        ('건물유형', basic.get('buildingType', ''), '시공부위', basic.get('constructionArea', '')),
        ('시공면적', f"{basic.get('constructionSize', '')}㎡" if basic.get('constructionSize') else '', '준공연도', basic.get('completionYear', '')),
        ('기존공법', basic.get('existingMethod', ''), '최초발생', basic.get('firstOccurrence', '')),
    ]
    for i, (k1, v1, k2, v2) in enumerate(fields):
        r = row + 1 + i
        _set_cell(ws, r, 1, '', NORMAL_FONT, CENTER)
        _set_cell(ws, r, 2, k1, HEADER_FONT, CENTER, LIGHT_FILL)
        ws.merge_cells(f'C{r}:D{r}')
        _set_cell(ws, r, 3, v1, NORMAL_FONT, CENTER)
        _set_cell(ws, r, 4, '', NORMAL_FONT, CENTER)
        _set_cell(ws, r, 5, k2, HEADER_FONT, CENTER, LIGHT_FILL)
        ws.merge_cells(f'F{r}:G{r}')
        _set_cell(ws, r, 6, v2, NORMAL_FONT, CENTER)
        _set_cell(ws, r, 7, '', NORMAL_FONT, CENTER)

    # 주요 증상
    row = 11
    ws.merge_cells(f'A{row}:G{row}')
    _set_cell(ws, row, 1, '■ 주요 증상', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    ws.merge_cells(f'A{row+1}:G{row+2}')
    _set_cell(ws, row+1, 1, basic.get('mainSymptoms', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 30
    ws.row_dimensions[row+2].height = 30
    _apply_border(ws, row+1, row+2, 1, 7)

    # 고객 정보
    row = 14
    ws.merge_cells(f'A{row}:G{row}')
    _set_cell(ws, row, 1, '■ 고객 정보', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    r = row + 1
    _set_cell(ws, r, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, r, 2, '고객명', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, r, 3, basic.get('clientName', ''), NORMAL_FONT, CENTER)
    _set_cell(ws, r, 4, '연락처', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{r}:G{r}')
    _set_cell(ws, r, 5, basic.get('clientPhone', ''), NORMAL_FONT, CENTER)
    for c in range(6, 8):
        ws.cell(row=r, column=c).border = BORDER_ALL

    r = row + 2
    _set_cell(ws, r, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, r, 2, '주소', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'C{r}:G{r}')
    _set_cell(ws, r, 3, basic.get('clientAddress', ''), NORMAL_FONT, WRAP)
    for c in range(4, 8):
        ws.cell(row=r, column=c).border = BORDER_ALL

    # 긴급도
    row = 17
    ws.merge_cells(f'A{row}:G{row}')
    _set_cell(ws, row, 1, '■ 긴급도 판단', SUBTITLE_FONT, Alignment(vertical='center'))
    urg_color = 'FCE4D6'
    if urgency.get('score', 0) >= 5:
        urg_color = 'FF9999'
    ws.cell(row=row, column=1).fill = PatternFill(start_color=urg_color, end_color=urg_color, fill_type='solid')

    r = row + 1
    _set_cell(ws, r, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, r, 2, '긴급도', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'C{r}:D{r}')
    _set_cell(ws, r, 3, urgency.get('level', ''), Font(name='맑은 고딕', bold=True, size=12), CENTER)
    _set_cell(ws, r, 5, '점수', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'F{r}:G{r}')
    _set_cell(ws, r, 6, f"{urgency.get('score', 0)}점 / 9점", NORMAL_FONT, CENTER)
    for c in [4, 7]:
        ws.cell(row=r, column=c).border = BORDER_ALL

    # 하단 서명란
    row = 20
    ws.merge_cells(f'A{row}:G{row}')
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    row = 21
    _set_cell(ws, row, 2, '작성자', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 5, '확인자', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 6, '', NORMAL_FONT, CENTER)
    row = 22
    ws.merge_cells(f'A{row}:G{row}')
    _set_cell(ws, row, 1, '(주)아이디에프이앤씨', NORMAL_FONT, CENTER)

    # 인쇄 설정
    ws.print_area = 'A1:G22'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════
# 2. 회의록
# ═══════════════════════════════════════════════

def create_meeting_minutes(data: dict, activity: dict, output_path: str) -> str:
    """회의록 엑셀을 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = '회의록'

    widths = [4, 14, 20, 14, 20, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    meta = data.get('metadata', {})
    minutes = activity.get('meetingMinutes', {})

    # 제목
    ws.merge_cells('A1:F1')
    _set_cell(ws, 1, 1, '(주)아이디에프이앤씨 회의록', TITLE_FONT, CENTER)
    ws.row_dimensions[1].height = 40

    # 기본 정보
    row = 3
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 2, 'AS 번호', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, meta.get('caseId', ''), NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '현장명', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    _set_cell(ws, row, 5, data.get('siteName', ''), NORMAL_FONT, CENTER)
    ws.cell(row=row, column=6).border = BORDER_ALL

    row = 4
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 2, '회의일시', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, f"{activity.get('date', '')} {activity.get('time', '')}", NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '장소', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    _set_cell(ws, row, 5, '현장', NORMAL_FONT, CENTER)
    ws.cell(row=row, column=6).border = BORDER_ALL

    # 참석자
    row = 6
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 참석자', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    attendees = minutes.get('attendees', [])
    row = 7
    for i, name in enumerate(attendees):
        col = (i % 3) * 2 + 2
        _set_cell(ws, row + i // 3, 1, '', NORMAL_FONT, CENTER)
        _set_cell(ws, row + i // 3, col, f'{i+1}', NORMAL_FONT, CENTER, LIGHT_FILL)
        _set_cell(ws, row + i // 3, col + 1, name, NORMAL_FONT, CENTER)
    att_rows = max(1, (len(attendees) + 2) // 3)
    _apply_border(ws, 7, 7 + att_rows - 1, 1, 6)

    # 회의 내용
    row = 7 + att_rows + 1
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 회의 내용/안건', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+2}')
    _set_cell(ws, row+1, 1, activity.get('content', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 40
    ws.row_dimensions[row+2].height = 40
    _apply_border(ws, row+1, row+2, 1, 6)

    # 결정사항
    row = row + 3
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 결정 사항', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+2}')
    _set_cell(ws, row+1, 1, minutes.get('decisions', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 35
    ws.row_dimensions[row+2].height = 35
    _apply_border(ws, row+1, row+2, 1, 6)

    # 보류사항
    row = row + 3
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 보류 사항', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+1}')
    _set_cell(ws, row+1, 1, minutes.get('pending', '없음'), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 30
    _apply_border(ws, row+1, row+1, 1, 6)

    # 액션 아이템
    row = row + 2
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 액션 아이템', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

    row += 1
    _set_cell(ws, row, 1, 'No', HEADER_FONT_W, CENTER, HEADER_FILL)
    ws.merge_cells(f'B{row}:C{row}')
    _set_cell(ws, row, 2, '할 일', HEADER_FONT_W, CENTER, HEADER_FILL)
    ws.cell(row=row, column=3).border = BORDER_ALL
    _set_cell(ws, row, 4, '담당자', HEADER_FONT_W, CENTER, HEADER_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    _set_cell(ws, row, 5, '기한', HEADER_FONT_W, CENTER, HEADER_FILL)
    ws.cell(row=row, column=6).border = BORDER_ALL

    items = minutes.get('actionItems', [])
    for i, item in enumerate(items):
        r = row + 1 + i
        _set_cell(ws, r, 1, str(i+1), NORMAL_FONT, CENTER)
        ws.merge_cells(f'B{r}:C{r}')
        _set_cell(ws, r, 2, item.get('task', ''), NORMAL_FONT, WRAP)
        ws.cell(row=r, column=3).border = BORDER_ALL
        _set_cell(ws, r, 4, item.get('assignee', ''), NORMAL_FONT, CENTER)
        ws.merge_cells(f'E{r}:F{r}')
        _set_cell(ws, r, 5, item.get('deadline', ''), NORMAL_FONT, CENTER)
        ws.cell(row=r, column=6).border = BORDER_ALL

    if not items:
        r = row + 1
        ws.merge_cells(f'A{r}:F{r}')
        _set_cell(ws, r, 1, '없음', NORMAL_FONT, CENTER)
        _apply_border(ws, r, r, 1, 6)

    # 서명란
    last_row = row + max(1, len(items)) + 2
    _set_cell(ws, last_row, 2, '작성자', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, last_row, 3, '', NORMAL_FONT, CENTER)
    _set_cell(ws, last_row, 4, '확인자', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, last_row, 5, '', NORMAL_FONT, CENTER)
    last_row += 1
    ws.merge_cells(f'A{last_row}:F{last_row}')
    _set_cell(ws, last_row, 1, '(주)아이디에프이앤씨', NORMAL_FONT, CENTER)

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════
# 3. 시공일지
# ═══════════════════════════════════════════════

def create_construction_log(data: dict, activity: dict, output_path: str) -> str:
    """시공일지 엑셀을 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = '시공일지'

    widths = [4, 14, 20, 14, 20, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    meta = data.get('metadata', {})
    log = activity.get('constructionLog', {})

    # 제목
    ws.merge_cells('A1:F1')
    _set_cell(ws, 1, 1, '(주)아이디에프이앤씨 시공일지', TITLE_FONT, CENTER)
    ws.row_dimensions[1].height = 40

    # 기본 정보
    row = 3
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 2, 'AS 번호', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, meta.get('caseId', ''), NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '현장명', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    _set_cell(ws, row, 5, data.get('siteName', ''), NORMAL_FONT, CENTER)
    ws.cell(row=row, column=6).border = BORDER_ALL

    row = 4
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 2, '시공일', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, activity.get('date', ''), NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '시공시간', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    _set_cell(ws, row, 5, activity.get('time', ''), NORMAL_FONT, CENTER)
    ws.cell(row=row, column=6).border = BORDER_ALL

    row = 5
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 2, '기상', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, log.get('weather', ''), NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '투입인원', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    _set_cell(ws, row, 5, f"{log.get('workers', '')}명", NORMAL_FONT, CENTER)
    ws.cell(row=row, column=6).border = BORDER_ALL

    # 시공 개요
    row = 7
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 시공 개요', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+1}')
    _set_cell(ws, row+1, 1, activity.get('content', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 30
    _apply_border(ws, row+1, row+1, 1, 6)

    # 작업 내용
    row = 9
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 작업 내용', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+3}')
    _set_cell(ws, row+1, 1, log.get('work', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 30
    ws.row_dimensions[row+2].height = 30
    ws.row_dimensions[row+3].height = 30
    _apply_border(ws, row+1, row+3, 1, 6)

    # 사용 자재
    row = 13
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 사용 자재', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+2}')
    _set_cell(ws, row+1, 1, log.get('materials', ''), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 30
    ws.row_dimensions[row+2].height = 30
    _apply_border(ws, row+1, row+2, 1, 6)

    # 진척도
    row = 16
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 진척도', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    row = 17
    _set_cell(ws, row, 1, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 2, '진척률', HEADER_FONT, CENTER, LIGHT_FILL)
    pct = log.get('progress', '0')
    _set_cell(ws, row, 3, f"{pct}%", Font(name='맑은 고딕', bold=True, size=14), CENTER)
    _set_cell(ws, row, 4, '상태', HEADER_FONT, CENTER, LIGHT_FILL)
    ws.merge_cells(f'E{row}:F{row}')
    status = '완료' if int(pct or 0) >= 100 else '진행중'
    _set_cell(ws, row, 5, status, NORMAL_FONT, CENTER)
    ws.cell(row=row, column=6).border = BORDER_ALL

    # 특이사항
    row = 19
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '■ 특이사항', SUBTITLE_FONT, Alignment(vertical='center'))
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws.merge_cells(f'A{row+1}:F{row+2}')
    _set_cell(ws, row+1, 1, log.get('notes', '없음'), NORMAL_FONT, WRAP)
    ws.row_dimensions[row+1].height = 30
    ws.row_dimensions[row+2].height = 30
    _apply_border(ws, row+1, row+2, 1, 6)

    # 서명란
    row = 22
    _set_cell(ws, row, 2, '현장소장', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 3, '', NORMAL_FONT, CENTER)
    _set_cell(ws, row, 4, '확인자', HEADER_FONT, CENTER, LIGHT_FILL)
    _set_cell(ws, row, 5, '', NORMAL_FONT, CENTER)
    row = 23
    ws.merge_cells(f'A{row}:F{row}')
    _set_cell(ws, row, 1, '(주)아이디에프이앤씨', NORMAL_FONT, CENTER)

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════
# 통합 생성: JSON → 모든 엑셀
# ═══════════════════════════════════════════════

def generate_all_excel(data: dict, output_dir: str) -> list:
    """AS 데이터에서 필요한 모든 엑셀 파일을 생성한다."""
    os.makedirs(output_dir, exist_ok=True)
    created = []
    case_id = data.get('metadata', {}).get('caseId', 'AS')
    safe_name = data.get('siteName', 'AS').replace(' ', '_')

    # 1. 공문접수서
    path = os.path.join(output_dir, f'공문접수서_{safe_name}_{case_id}.xlsx')
    create_document_receipt(data, path)
    created.append(path)
    print(f'  ✅ 공문접수서: {path}')

    # 2. 회의록 (활동 중 meeting 타입)
    for act in data.get('activities', []):
        if act.get('type') == 'meeting':
            seq = data['activities'].index(act) + 1
            path = os.path.join(output_dir, f'회의록_{seq}차_{safe_name}_{case_id}.xlsx')
            create_meeting_minutes(data, act, path)
            created.append(path)
            print(f'  ✅ 회의록 {seq}차: {path}')

    # 3. 시공일지 (활동 중 construction 타입)
    day = 0
    for act in data.get('activities', []):
        if act.get('type') == 'construction':
            day += 1
            path = os.path.join(output_dir, f'시공일지_{day}일차_{safe_name}_{case_id}.xlsx')
            create_construction_log(data, act, path)
            created.append(path)
            print(f'  ✅ 시공일지 {day}일차: {path}')

    return created


# ──── CLI 실행 ────
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('사용법: python create_excel.py <json_file> [output_dir]')
        print('  json_file: AS 데이터 JSON 파일')
        print('  output_dir: 출력 폴더 (기본: json 파일과 같은 폴더)')
        sys.exit(1)

    json_path = sys.argv[1]
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(json_path) or '.'

    print(f'\n📊 엑셀 양식 생성 시작...')
    print(f'   입력: {json_path}')
    print(f'   출력: {output_dir}\n')

    files = generate_all_excel(data, output_dir)
    print(f'\n✅ 완료! 총 {len(files)}개 파일 생성')
